# -*- coding: utf-8 -*-
"""
Created on Mon Sep 26 15:20:04 2016

@author: zzpp220
"""

from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws.title="chart"
b4 = ws['B4']
ws['B4'] = 4444
#==============================================================================
# tuple(ws.iter_rows("A1:C3"))
# ((<Cell python.A1>, <Cell Python.B1>, <Cell Python.C1>), 
#  (<Cell python.A2>, <Cell Python.B2>, <Cell Python.C2>), 
#  (<Cell python.A3>, <Cell Python.B3>, <Cell Python.C3>))
#==============================================================================
